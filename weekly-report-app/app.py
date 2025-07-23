import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import os

# === CONFIG ===
EXCEL_PATH = https://racetrac.sharepoint.com/sites/EngineeringLeadership/Shared%20Documents/General/Larry%20Sloan/Construction%20Weekly%20Updates/Construction%20Weekly%20Updates.xlsx
SHEET_NAME = "Sheet1"

# === WEEK LABEL ===
today = date.today()
week_label = f"Week {today.isocalendar()[1]} {today.year}"

st.title("🧱 Construction Milestone Update")

# === FORM ===
with st.form("milestone_form", clear_on_submit=True):
    st.write(f"### Submit Milestone for {week_label}")

    project_name = st.text_input("Project Name")
    milestone = st.selectbox("Milestone", ["Store Opening", "TCO", "Turnover", "Permit", "Equipment Set", "Other"])
    baseline_date = st.date_input("Baseline Date")
    last_week_date = st.date_input("Last Week Date")
    this_week_date = st.date_input("This Week Date")
    notes = st.text_area("Notes (each bullet on new line)")

    submitted = st.form_submit_button("Submit")

# === SUBMIT HANDLER ===
if submitted:
    # Calculate delta
    delta_days = (this_week_date - baseline_date).days
    trend = "📈 Improved" if delta_days < 0 else "📉 Slipped" if delta_days > 0 else "➖ No Change"

    # Convert notes to clean text (bulleted for form only)
    notes_clean = notes.strip().replace("\n", " • ")

    # New data row
    new_data = {
        "Week": week_label,
        "Project Name": project_name,
        "Milestone": milestone,
        "Baseline Date": baseline_date.strftime("%Y-%m-%d"),
        "Last Week Date": last_week_date.strftime("%Y-%m-%d"),
        "This Week Date": this_week_date.strftime("%Y-%m-%d"),
        "Δ Days": f"{delta_days:+}",
        "Trend": trend,
        "Notes": notes_clean
    }
    new_row = pd.DataFrame([new_data])

    # === Append to Excel ===
    try:
        if os.path.exists(EXCEL_PATH):
            book = load_workbook(EXCEL_PATH)
            writer = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets[SHEET_NAME].max_row
            new_row.to_excel(writer, sheet_name=SHEET_NAME, startrow=startrow, index=False, header=False)
            writer.save()
            writer.close()
        else:
            new_row.to_excel(EXCEL_PATH, index=False)

        st.success(f"✅ Entry saved for {project_name} ({milestone})")
    except Exception as e:
        st.error(f"❌ Failed to save entry: {e}")
